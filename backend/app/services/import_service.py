import io
import re
import csv
import datetime as dt
from typing import List, Dict, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import openpyxl

from app.schemas.import_schema import ImportedRowSchema, ImportPreviewResponse, ImportExecuteResponse
from app.models.models import Especime, Taxonomia, LocalidadeGeografica, StatusEspecime, TipoColeta


def _slug(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    return re.sub(r"\s+", " ", str(text)).strip() or None


def _clean_name(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    v = re.sub(r"\s+", " ", str(text)).strip()
    return v or None


def _parse_date(value: object) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if not s or s.lower() in {"desconhecido", "desconhecida", "-", "n/d"}:
        return None
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d", "%d\\%m\\%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _detect_and_map_headers(header_tuple: tuple) -> Dict[int, str]:
    header = tuple((v or "").strip().lower() for v in header_tuple)
    alias_map = {
        "nome cientifico": "nome_cientifico",
        "nome científico": "nome_cientifico",
        "nome popular": "nome_popular",
        "familia": "familia",
        "família": "familia",
        "genero": "genero",
        "gênero": "genero",
        "especie": "especie",
        "espécie": "especie",
        "localizacao": "localizacao",
        "localização": "localizacao",
        "data da coleta": "data_coleta",
        "data de coleta": "data_coleta",
        "data coleta": "data_coleta",
        "origem": "origem",
        "coletor": "coletor",
        "codigo": "codigo",
        "código": "codigo",
        "filo": "filo",
        "classe": "classe",
        "ordem": "ordem",
    }
    result = {}
    for idx, h in enumerate(header):
        mapped = alias_map.get(h, h)
        if mapped in alias_map.values():
            result[idx] = mapped
    return result


class ImportService:

    @staticmethod
    def parse_and_validate_file(file_bytes: bytes, filename: str) -> ImportPreviewResponse:
        raw_rows = []

        if filename.lower().endswith(".csv"):
            text = file_bytes.decode("utf-8-sig", errors="ignore")
            lines = text.splitlines()
            reader = csv.reader(lines)
            headers = None
            for row_idx, row in enumerate(reader, start=1):
                if not row or not any(row):
                    continue
                if headers is None:
                    headers = _detect_and_map_headers(tuple(row))
                    if not headers:
                        headers = {i: f"col_{i}" for i in range(len(row))}
                    continue

                raw_rows.append((row_idx, "CSV", row, headers))
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = None
                for row_idx, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
                    if not row_values or not any(v is not None for v in row_values):
                        continue
                    str_row = tuple("" if v is None else str(v) for v in row_values)
                    joined = " ".join(str_row).upper()
                    if any(token in joined for token in ["CÓDIGO", "CODIGO", "NOME CIENTIFICO", "FAMILIA"]):
                        headers = _detect_and_map_headers(str_row)
                        continue
                    if headers is None:
                        continue
                    raw_rows.append((row_idx, sheet_name, row_values, headers))

        parsed_rows: List[ImportedRowSchema] = []
        seen_codes = set()
        seen_taxonomias = set()
        seen_localidades = set()

        valid_cnt = 0
        warn_cnt = 0
        err_cnt = 0

        for idx, (row_idx, sheet, row, headers) in enumerate(raw_rows, start=1):
            def get_val(field_name: str) -> Optional[str]:
                for h_idx, mapped in headers.items():
                    if mapped == field_name and h_idx < len(row):
                        val = row[h_idx]
                        if val is not None:
                            return _clean_name(str(val))
                return None

            codigo = get_val("codigo") or f"HERB-AUTO-{idx:04d}"
            nome_cientifico = get_val("nome_cientifico") or get_val("especie") or get_val("genero") or "Escipie não identificada"
            nome_popular = get_val("nome_popular")
            filo = get_val("filo")
            classe = get_val("classe")
            ordem = get_val("ordem")
            familia = get_val("familia")
            genero = get_val("genero")
            especie = get_val("especie")
            localizacao = get_val("localizacao")
            raw_date = get_val("data_coleta")
            origem = get_val("origem")
            coletor = get_val("coletor")

            parsed_date = _parse_date(raw_date)
            data_coleta_str = parsed_date.isoformat() if parsed_date else None

            status_val = "valid"
            msg_val = None

            if codigo in seen_codes:
                status_val = "warning"
                msg_val = f"Código {codigo} duplicado na planilha. Será atualizado se existir no banco."
                warn_cnt += 1
            else:
                seen_codes.add(codigo)
                valid_cnt += 1

            if nome_cientifico:
                seen_taxonomias.add(nome_cientifico.lower())
            if localizacao:
                seen_localidades.add(localizacao.lower())

            parsed_rows.append(
                ImportedRowSchema(
                    sheet=sheet,
                    row_index=row_idx,
                    codigo=codigo,
                    nome_cientifico=nome_cientifico,
                    nome_popular=nome_popular,
                    filo=filo,
                    classe=classe,
                    ordem=ordem,
                    familia=familia,
                    genero=genero,
                    especie=especie,
                    localizacao=localizacao,
                    data_coleta=data_coleta_str,
                    origem=origem,
                    coletor=coletor,
                    status_validacao=status_val,
                    mensagem_validacao=msg_val,
                )
            )

        return ImportPreviewResponse(
            filename=filename,
            total_rows=len(parsed_rows),
            valid_rows=valid_cnt,
            warning_rows=warn_cnt,
            error_rows=err_cnt,
            new_taxonomias_est=len(seen_taxonomias),
            new_localidades_est=len(seen_localidades),
            rows=parsed_rows,
        )

    @staticmethod
    async def execute_import(db: AsyncSession, rows: List[ImportedRowSchema], user_id: int) -> ImportExecuteResponse:
        especimes_criados = 0
        especimes_atualizados = 0
        taxonomias_criadas = 0
        localidades_criadas = 0
        erros = 0
        detalhes_erros = []

        tax_res = await db.execute(select(Taxonomia))
        all_tax = tax_res.scalars().all()
        tax_cache: Dict[str, int] = {t.nome_cientifico.lower(): t.id for t in all_tax}

        loc_res = await db.execute(select(LocalidadeGeografica))
        all_loc = loc_res.scalars().all()
        loc_cache: Dict[str, int] = {}
        for l in all_loc:
            key = (l.localidade or "").strip().lower()
            if key:
                loc_cache[key] = l.id

        for r in rows:
            try:
                # 1. Process Taxonomia
                tax_name = r.nome_cientifico.strip() if r.nome_cientifico else "Não identificada"
                tax_key = tax_name.lower()

                tax_id = tax_cache.get(tax_key)
                if not tax_id:
                    new_tax = Taxonomia(
                        nome_cientifico=tax_name,
                        familia=r.familia,
                        genero=r.genero or (tax_name.split()[0] if " " in tax_name else tax_name),
                        epiteto_especifico=r.especie,
                        nome_comum=r.nome_popular,
                        filo=r.filo,
                        classe=r.classe,
                        ordem=r.ordem,
                    )
                    db.add(new_tax)
                    await db.flush()
                    tax_id = new_tax.id
                    tax_cache[tax_key] = tax_id
                    taxonomias_criadas += 1

                # 2. Process Localidade
                loc_id = None
                if r.localizacao:
                    loc_key = r.localizacao.strip().lower()
                    loc_id = loc_cache.get(loc_key)
                    if not loc_id:
                        new_loc = LocalidadeGeografica(
                            pais="Brasil",
                            localidade=r.localizacao,
                            datum_geodesico="WGS84",
                            criado_em=dt.datetime.utcnow(),
                        )
                        db.add(new_loc)
                        await db.flush()
                        loc_id = new_loc.id
                        loc_cache[loc_key] = loc_id
                        localidades_criadas += 1

                # 3. Upsert Especime
                esp_res = await db.execute(select(Especime).where(Especime.codigo_catalogo == r.codigo))
                esp = esp_res.scalar_one_or_none()

                parsed_dt = None
                if r.data_coleta:
                    try:
                        parsed_dt = dt.datetime.fromisoformat(r.data_coleta)
                    except ValueError:
                        pass

                if esp:
                    esp.taxonomia_id = tax_id
                    if loc_id:
                        esp.localidade_id = loc_id
                    if r.coletor:
                        esp.coletor_principal = r.coletor
                    if parsed_dt:
                        esp.data_coleta = parsed_dt
                    especimes_atualizados += 1
                else:
                    esp = Especime(
                        codigo_catalogo=r.codigo,
                        taxonomia_id=tax_id,
                        localidade_id=loc_id,
                        coletor_principal=r.coletor,
                        data_coleta=parsed_dt,
                        status=StatusEspecime.ativo,
                        tipo_coleta=TipoColeta.campo,
                        numero_individuos=1,
                        direitos="CC BY 4.0",
                        data_entrada_colecao=dt.datetime.utcnow(),
                    )
                    db.add(esp)
                    especimes_criados += 1

            except Exception as exc:
                erros += 1
                detalhes_erros.append(f"Linha {r.row_index} ({r.codigo}): {str(exc)}")

        await db.commit()

        return ImportExecuteResponse(
            total_processed=len(rows),
            especimes_criados=especimes_criados,
            especimes_atualizados=especimes_atualizados,
            taxonomias_criadas=taxonomias_criadas,
            localidades_criadas=localidades_criadas,
            erros=erros,
            detalhes_erros=detalhes_erros,
        )
