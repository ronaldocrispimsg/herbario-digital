#!/usr/bin/env python3
"""Importador do acervo do Herbário Digital a partir de 'Organização Herbário.xlsx'.

Modos:
  dry-run  -> apenas abre a planilha e gera um relatório estruturado em stdout/JSON.
  import   -> insere/atualiza via API REST do BioAcervo.

Uso:
  python scripts/import_herbario.py --mode dry-run
  python scripts/import_herbario.py --mode import --url http://localhost:8000 --token <JWT> --admin-user-id 1
  python scripts/import_herbario.py --mode import --url http://localhost:8000 --login admin@example.com --password admin123 --admin-user-id 1

Projeto: ~/Projetos/herbario-digital
Design:
- Paralelismo controlado por semáforo asyncio
- Idempotência por codigo_catalog
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import datetime as dt
import hashlib
import json
import math
import os
import re
import sys
import urllib.parse
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable, Optional

try:
    import httpx
except Exception:
    httpx = None  # type: ignore

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl necessária: {exc}")

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = PROJECT_ROOT / "Organização Herbário.xlsx"
REPORT_PATH = PROJECT_ROOT / "scripts" / "import_report.json"
MAX_CONCURRENT = 8  # estabilidade para teste/prod leve


# ── Modelos de registro bruto e normalizado ──────────────────────────────────

@dataclass
class RawRow:
    sheet: str
    row_index: int
    codigo: str | None
    nome_cientifico: str | None
    nome_popular: str | None
    filo: str | None
    classe: str | None
    ordem: str | None
    familia: str | None
    genero: str | None
    especie: str | None
    localizacao: str | None
    data_coleta_raw: object
    origem: str | None
    coletor: str | None


@dataclass
class ImportedRow:
    sheet: str
    row_index: int
    codigo: str
    nome_cientifico: str
    nome_popular: str | None = None
    filo: str | None = None
    classe: str | None = None
    ordem: str | None = None
    familia: str | None = None
    genero: str | None = None
    especie: str | None = None
    localizacao: str | None = None
    data_coleta: str | None = None
    origem: str | None = None
    coletor: str | None = None

    taxonomia_id: int | None = None
    localidade_id: int | None = None
    especie_id: int | None = None
    status: str | None = None
    error: str | None = None
    action: str | None = None


# ── Parsing da planilha ──────────────────────────────────────────────────────

def _slug(text: str | None) -> str | None:
    if not text:
        return None
    return re.sub(r"\s+", " ", str(text)).strip() or None


def _clean_name(text: str | None) -> str | None:
    if not text:
        return None
    v = re.sub(r"\s+", " ", str(text)).strip()
    return v or None


def _parse_date(value: object) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if not s or s.lower() in {"desconhecido", "desconhecida", "-", "n/d"}:
        return None
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d", "%d\\%m\\%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _detect_and_map_headers(values: tuple) -> dict[int, str] | None:
    header = tuple((v or "").strip().lower() for v in values)
    candidates = [
        ("codigo", "nome cientifico", "nome popular", "filo", "classe", "ordem", "familia", "genero", "especie", "origem", "coletor", "data da coleta", "localizacao"),
        ("codigo", "nome cientifico", "nome popular", "filo", "classe", "ordem", "familia", "genero", "especie", "origem", "coletor", "data de coleta", "localizacao"),
        ("codigo", "nome cientifico", "nome popular", "filo", "classe", "ordem", "familia", "genero", "especie", "localizacao", "data da coleta", "origem", "coletor"),
        ("codigo", "nome cientifico", "nome popular", "filo", "classe", "ordem", "familia", "genero", "especie", "localizacao", "data de coleta", "origem", "coletor"),
    ]

    alias_map = {
        "nome cientifico": "nome_cientifico",
        "nome popular": "nome_popular",
        "familia": "familia",
        "familía": "familia",
        "genero": "genero",
        "género": "genero",
        "especie": "especie",
        "espécie": "especie",
        "localizacao": "localizacao",
        "localização": "localizacao",
        "data da coleta": "data_coleta",
        "data de coleta": "data_coleta",
        "data coleta": "data_coleta",
        "origem": "origem",
        "coletor": "coletor",
        "codigo": "codigo",
        "código": "codigo",
        "filo": "filo",
        "classe": "classe",
        "ordem": "ordem",
    }

    normalized = []
    for h in header:
        normalized.append(alias_map.get(h, h))
    for candidate in candidates:
        if all(field in normalized for field in candidate):
            return {i: normalized[i] for i in range(len(normalized))}
    return {i: normalized[i] for i in range(len(normalized)) if normalized[i]}


def _coerce(row: tuple, headers: dict[int, str]) -> RawRow | None:
    def get(field_name: str) -> str | None:
        for idx, mapped in headers.items():
            if mapped == field_name:
                value = row[idx] if idx < len(row) else None
                if value is None:
                    return None
                return _clean_name(str(value))
        return None

    codigo = get("codigo")
    nome_cientifico = get("nome_cientifico")
    if not codigo and not nome_cientifico:
        return None

    return RawRow(
        sheet="",  # filled later
        row_index=0,  # filled later
        codigo=codigo,
        nome_cientifico=nome_cientifico,
        nome_popular=get("nome_popular"),
        filo=get("filo"),
        classe=get("classe"),
        ordem=get("ordem"),
        familia=get("familia"),
        genero=get("genero"),
        especie=get("especie"),
        localizacao=get("localizacao"),
        data_coleta_raw=None,
        origem=get("origem"),
        coletor=get("coletor"),
    )


def parse_xlsx(path: Path) -> tuple[list[RawRow], dict[str, int]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[RawRow] = []
    stats: dict[str, int] = {"sheets": len(wb.sheetnames), "parsed": 0, "skipped": 0}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            text_row = tuple("" if v is None else str(v) for v in row)
            joined = " ".join(text_row).upper()
            if any(token in joined for token in ["CÓDIGO", "CODIGO", "CÓDIGO "]):
                headers = _detect_and_map_headers(text_row)
                continue
            if headers is None:
                continue
            if any(v is None for v in row[: min(8, len(row))]):
                if all(v is None for v in row):
                    continue
            if not any(v is not None for v in row):
                continue
            parsed = _coerce(row, headers)
            if parsed:
                parsed.sheet = sheet_name
                parsed.row_index = 0  # kept externally later
                rows.append(parsed)
                stats["parsed"] += 1
            else:
                stats["skipped"] += 1

    for idx, r in enumerate(rows, start=1):
        r.row_index = idx

    return rows, stats


# ── Normalização / dedup / anonimização leve ────────────────────────────────

def normalize_rows(rows: Iterable[RawRow]) -> list[ImportedRow]:
    normalized: list[ImportedRow] = []
    seen_codes: set[str] = set()
    seen_composite: set[str] = set()
    for r in rows:
        code = _slug(r.codigo) or f"ROW-{r.row_index}"
        nome_cientifico = _clean_name(r.nome_cientifico)
        nome_popular = _clean_name(r.nome_popular)
        coletor = _clean_name(r.coletor)
        localizacao = _clean_name(r.localizacao)
        origem = _clean_name(r.origem)
        if not nome_cientifico and not localizacao and not coletor:
            continue
        composite = "|".join(x or "" for x in [code, nome_cientifico or "", (r.familia or "").strip(), (r.data_coleta_raw or "").strip()])
        if composite in seen_composite:
            continue
        seen_composite.add(composite)
        if code in {"desconhecido", "desconhecida"}:
            code = f"ROW-{r.row_index}"
        if code in seen_codes and not nome_cientifico:
            continue
        seen_codes.add(code)
        parsed_date = _parse_date(r.data_coleta_raw)
        data_coleta_str = parsed_date.isoformat() if parsed_date else None
        normalized.append(
            ImportedRow(
                sheet=r.sheet,
                row_index=r.row_index,
                codigo=code,
                nome_cientifico=nome_cientifico or code,
                nome_popular=nome_popular,
                filo=_clean_name(r.filo),
                classe=_clean_name(r.classe),
                ordem=_clean_name(r.ordem),
                familia=_clean_name(r.familia),
                genero=_clean_name(r.genero),
                especie=_clean_name(r.especie),
                localizacao=localizacao,
                data_coleta=data_coleta_str,
                origem=origem,
                coletor=coletor,
            )
        )
    return normalized


# ── Relatórios/exportação para revisão ──────────────────────────────────────

def export_report(rows: list[ImportedRow], path: Path) -> None:
    payload = {
        "generated_at": dt.datetime.utcnow().isoformat() + "Z",
        "total": len(rows),
        "with_date": sum(1 for r in rows if r.data_coleta),
        "without_date": sum(1 for r in rows if not r.data_coleta),
        "rows": [asdict(r) for r in rows[:5000]],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_csv(rows: list[ImportedRow], path: Path) -> None:
    fields = [
        "sheet", "row_index", "codigo", "nome_cientifico", "nome_popular", "filo", "classe", "ordem", "familia",
        "genero", "especie", "localizacao", "data_coleta", "origem", "coletor", "taxonomia_id", "localidade_id",
        "especie_id", "status", "error", "action",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow(asdict(r))


# ── API client ───────────────────────────────────────────────────────────────

class HerbarioAPI:
    def __init__(
        self,
        base_url: str,
        *,
        token: str | None = None,
        login: str | None = None,
        password: str | None = None,
        admin_user_id: int = 1,
        semaphore: asyncio.Semaphore | None = None,
        client: httpx.AsyncClient | None = None,
        dry_run: bool = False,
    ) -> None:
        if httpx is None:
            raise SystemExit("httpx não instalado no ambiente do script")
        self.login = login
        self.password = password
        self.base_url = base_url.rstrip("/")
        self.admin_user_id = admin_user_id
        self.dry_run = dry_run
        self.token = token
        self.headers: dict[str, str] = {"accept": "application/json"}
        self.semaphore = semaphore or asyncio.Semaphore(MAX_CONCURRENT)
        self._client = client or httpx.AsyncClient(timeout=httpx.Timeout(connect=10.0, read=30.0, write=30.0, pool=5.0))

    async def _obter_token(self) -> str:
        if self.token:
            return self.token
        payload = {"username": self.login, "password": self.password}
        body = urllib.parse.urlencode(payload)
        r = await self._client.post(
            f"{self.base_url}/api/v1/auth/login",
            content=body,
            headers={"accept": "application/json", "content-type": "application/x-www-form-urlencoded"},
        )
        r.raise_for_status()
        data = r.json()
        self.token = data["access_token"]
        return self.token

    async def _request(self, method: str, url: str, **kwargs):
        token = await self._obter_token()
        headers = kwargs.pop("headers", {})
        headers["authorization"] = f"Bearer {token}"
        full = url if url.startswith("http") else f"{self.base_url}{url}"
        async with self.semaphore:
            return await self._client.request(method, full, headers=headers, **kwargs)

    async def _get(self, url: str, **kwargs):
        return await self._request("GET", url, **kwargs)

    async def _post(self, url: str, **kwargs):
        return await self._request("POST", url, **kwargs)

    async def _put(self, url: str, **kwargs):
        return await self._request("PUT", url, **kwargs)

    async def _paginated_get(self, url: str, *, params: dict | None = None, limit: int = 100) -> list[dict]:
        page = 1
        out: list[dict] = []
        params = params or {}
        params.setdefault("per_page", limit)
        while True:
            params["page"] = page
            r = await self._get(url, params=params)
            r.raise_for_status()
            data = r.json()
            items = data.get("items", []) if isinstance(data, dict) else data
            out.extend(items)
            if not items or page * limit >= (data.get("total", 0) if isinstance(data, dict) else len(items)):
                break
            page += 1
        return out

    async def fetch_taxonomias(self) -> list[dict]:
        return await self._paginated_get("/api/v1/taxonomias")

    async def fetch_localidades(self) -> list[dict]:
        return await self._paginated_get("/api/v1/localidades")

    async def create_taxonomia(self, body: dict) -> dict:
        if self.dry_run:
            return {"id": 0, **body}
        r = await self._post("/api/v1/taxonomias", json=body)
        r.raise_for_status()
        return r.json()

    async def create_especime(self, body: dict) -> dict:
        if self.dry_run:
            return {"id": 0, **body}
        r = await self._post("/api/v1/especimes", json=body)
        r.raise_for_status()
        return r.json()

    async def create_localidade(self, body: dict) -> dict:
        if self.dry_run:
            return {"id": 0, **body}
        r = await self._post("/api/v1/localidades", json=body)
        r.raise_for_status()
        return r.json()

    async def close(self) -> None:
        await self._client.aclose()


# ── Caso de importação por item ─────────────────────────────────────────────

def _taxonomia_body(r: ImportedRow) -> dict:
    return {
        "nome_cientifico": r.nome_cientifico,
        "familia": r.familia,
        "genero": r.genero,
        "reino": r.filo,
        "classe": r.classe,
        "ordem": r.ordem,
        "autor_descricao": None,
        "ano_descricao": None,
        "nome_comum": r.nome_popular,
        "sinonimos": None,
        "notas_taxonomicas": None,
    }


def _localidade_body(r: ImportedRow) -> dict | None:
    if not r.localizacao:
        return None
    return {
        "pais": "Brasil",
        "estado": _extract_uf(r.localizacao),
        "municipio": _extract_municipio(r.localizacao),
        "localidade": r.localizacao,
        "latitude": None,
        "longitude": None,
        "altitude_m": None,
        "datum_geodesico": "WGS84",
        "precisao_coordenadas_m": None,
        "metodo_geolocalizacao": None,
        "bioma": None,
    }


def _especime_body(r: ImportedRow, taxonomia_id: int | None, localidade_id: int | None) -> dict:
    data_coleta = dt.date.fromisoformat(r.data_coleta) if r.data_coleta else None
    return {
        "codigo_catalogo": r.codigo,
        "taxonomia_id": taxonomia_id,
        "data_coleta": data_coleta.isoformat() if data_coleta else None,
        "tipo_coleta": "campo",
        "coletor_principal": r.coletor,
        "coletores_adicionais": None,
        "localidade_id": localidade_id,
        "sexo": None,
        "estagio_vida": None,
        "condicao": None,
        "numero_individuos": 1,
        "descricao_morfologica": None,
        "observacoes": None,
        "habitat": None,
        "identificado_por": None,
        "data_identificacao": None,
        "metodo_identificacao": None,
        "nivel_confianca_id": "média",
        "status": "ativo",
        "localizacao_fisica": r.localizacao,
        "meio_preservacao": "seco",
        "dwc_dataset_id": None,
        "referencias_bibliograficas": None,
        "direitos": "CC BY 4.0",
        "licenca": None,
    }


_UF_RE = re.compile(r"\b([A-Z]{2})\b")
_MUN_RE = re.compile(r"([A-Za-zÀ-ÖØ-öø-ÿ\s]+(?:Januária|Bonito de Minas|Pedras de Maria da Cruz|Cônego Marinho|Itacarambi|Patis|Lontra|São Francisco|Juvenília|Juvenilha|Fazenda|Comunidade|Centro|Brejo do amparo|Olhos d'água|Vila Jadete|Barreiro|Itacaranbi|Itacarambi|Projeto Jaíba|Barão de São Romão))\s*[-/,]?\s*(?:[-–]\s*)?(?:MG)?", re.IGNORECASE)


def _extract_uf(text: str) -> str | None:
    m = _UF_RE.search(text)
    if m:
        return m.group(1).upper()
    return "MG" if re.search(r"Januaria|Minas|MG|Cônego Marinho", text, re.IGNORECASE) else None


def _extract_municipio(text: str) -> str | None:
    m = _MUN_RE.search(text)
    if m:
        return m.group(1).strip()
    tokens = re.split(r"\s*[-/,]\s*", text)
    if tokens:
        return tokens[0].strip()
    return None


# ── Pipeline de importação ───────────────────────────────────────────────────

async def run_import(api: HerbarioAPI, rows: list[ImportedRow]) -> list[ImportedRow]:
    if api.dry_run:
        out: list[ImportedRow] = []
        for r in rows:
            r.action = "would_create"
            out.append(r)
        return out

    out: list[ImportedRow] = []
    for r in rows:
        r.action = "error"
        r.error = None
        try:
            taxa_body = _taxonomia_body(r)
            taxa = None
            try:
                taxa = await api.create_taxonomia(taxa_body)
            except httpx.HTTPStatusError as exc:
                if exc.response.status_code == 409:
                    taxa = {"id": None, "nome_cientifico": taxa_body.get("nome_cientifico")}
                else:
                    r.error = f"taxonomia {exc.response.status_code}: {exc.response.text}"
                    out.append(r)
                    continue
            tax_id = taxa.get("id") if isinstance(taxa, dict) else None
            r.taxonomia_id = tax_id
            if tax_id is None:
                r.error = "taxonomia nao retornou id"
                out.append(r)
                continue

            loc_body = _localidade_body(r)
            loc_id = None
            if loc_body:
                try:
                    loc = await api.create_localidade(loc_body)
                    loc_id = loc.get("id") if isinstance(loc, dict) else None
                except httpx.HTTPStatusError as exc:
                    if exc.response.status_code == 409:
                        loc_id = None
                    else:
                        r.error = f"localidade {exc.response.status_code}: {exc.response.text}"
                        out.append(r)
                        continue
                except Exception:
                    loc_id = None
            r.localidade_id = loc_id

            esp_body = _especime_body(r, tax_id, loc_id)
            if loc_id is None:
                esp_body.pop("localidade_id", None)
            created = await api.create_especime(esp_body)
            r.especime_id = created.get("id")
            r.action = "created"
        except httpx.HTTPStatusError as exc:
            r.error = f"{exc.response.status_code}: {exc.response.text}"
        except Exception as exc:
            r.error = repr(exc)
        out.append(r)
    return out


# ── Main CLI ─────────────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(description="Importador do acervo do Herbário Digital")
    p.add_argument("--mode", choices=["dry-run", "import"], default="dry-run")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--url", default="http://localhost:8000")
    p.add_argument("--token")
    p.add_argument("--login")
    p.add_argument("--password")
    p.add_argument("--admin-user-id", type=int, default=1)
    p.add_argument("--report", type=Path, default=REPORT_PATH)
    p.add_argument("--csv", type=Path, default=REPORT_PATH.with_suffix(".csv"))
    args = p.parse_args()

    if not args.xlsx.exists():
        print(f"Arquivo não encontrado: {args.xlsx}")
        return 2

    print(f"✅ Abrindo planilha: {args.xlsx}")
    raw_rows, stats = parse_xlsx(args.xlsx)
    print(f"Planilhas={stats['sheets']} parsed={stats['parsed']} skipped={stats['skipped']}")

    normalized = normalize_rows(raw_rows)
    print(f"Normalizados={len(normalized)}")

    async def _execute():
        api = HerbarioAPI(
            args.url,
            token=args.token,
            login=args.login,
            password=args.password,
            admin_user_id=args.admin_user_id,
            dry_run=args.mode == "dry-run",
        )
        try:
            results = await run_import(api, normalized)
        finally:
            await api.close()
        return results

    results = asyncio.run(_execute())

    export_report(results, args.report)
    export_csv(results, args.csv)
    print(f"Relatório JSON: {args.report}")
    print(f"Relatório CSV : {args.csv}")
    created = sum(1 for r in results if r.action == "created")
    would_create = sum(1 for r in results if r.action == "would_create")
    errors = sum(1 for r in results if r.action == "error")
    print(f"MODO={args.mode} criados={created} simulação={would_create} erros={errors}")
    if errors:
        sample = [r for r in results if r.action == "error"][:10]
        for r in sample:
            print(f"- {r.sheet}:{r.row_index} {r.codigo} => {r.error}")
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
